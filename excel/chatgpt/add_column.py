import openpyxl

# 打开Excel文件
workbook = openpyxl.load_workbook('data/成绩表.xlsx')

# 获取第一个工作表
sheet = workbook.active

# 添加表头
sheet.cell(row=1, column=6, value='平均分')

# 计算平均分
for i in range(2, sheet.max_row + 1):
    score_total = 0
    count = 0
    for j in range(3, 6):
        value = sheet.cell(row=i, column=j).value
        if isinstance(value, str) and not value.isnumeric():
            continue
        score_total += float(value)
        count += 1
    avg_score = score_total / count if count > 0 else 0
    sheet.cell(row=i, column=6, value=avg_score)


# 保存修改后的Excel文件
workbook.save('Table1.xlsx')
